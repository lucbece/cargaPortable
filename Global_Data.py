import pandas as pd
import os
import shutil
from openpyxl import Workbook, load_workbook 

rutaOUT = 'D:/'

def initialize():
    global empresas_list
    f = open(str(os.path.dirname(os.path.realpath(__file__))) + '/database/Empresas.txt','r')
    empresas_list = f.read().split(',')
    empresas_list.sort()
    f.close()
    return empresas_list

def add_empresa(new_empresa):
    f = open(str(os.path.dirname(os.path.realpath(__file__))) + '/database/Empresas.txt','a')
    f.write(',' + new_empresa)
    f.close()
    txt = open(str(os.path.dirname(os.path.realpath(__file__))) + '/database/registro/' + new_empresa + '.txt', 'a+')
    txt.close()
def pick_database(value):
    Empresa = empresas_list[value]
    df = pd.read_excel (os.path.dirname(os.path.realpath(__file__)) + '/database/database/' + Empresa + '.xls')
    Equipos = df['Equipo']
    Componente = df['Componente']
    PE = df['P.E']
    returnable = []
    for i in range(0,len(Equipos)):
        if PE[i] == 'General' or PE[i] == 'Sin Especificar':
            returnable.append(str(Equipos[i]) + ' ' + str(Componente[i]))
        else:
            returnable.append(str(Equipos[i]) + ' ' + str(Componente[i]) + ' ' + str(PE[i]))
    return returnable

def numerar_list(raw_list, ordered_list):
    int_list = []
    for muestra  in raw_list:
        for i in range (len(raw_list)):
            if muestra == ordered_list[i][0]:
                int_list.append(i+1)
    return int_list

def EXCEL_carga(Empresa, CDM, Muestras_list, Fecha):
    if 'YPF' in Empresa:
        file = os.path.dirname(os.path.realpath(__file__)) + '/Formularios/F-GA-9_2 Analisis de lubricantes_ypf.xlsx'
    else:
        file = os.path.dirname(os.path.realpath(__file__)) + "/Formularios/F-GA-9_2 Analisis de lubricantes.xlsx"

    File_name = Fecha + '.xlsx'    

    if not os.path.exists(rutaOUT + "Para carga/Midiendo/"+ Empresa) == True:
        os.makedirs(rutaOUT + "Para carga/Midiendo/"+ Empresa)
    shutil.copy(file, File_name)


    salida = load_workbook(File_name)
    sheet = salida.active

    sheet.cell(row=6,column=2).value = Empresa
    sheet.cell(row=6, column=5).value = Fecha
    sheet.cell(row=6, column=7).value = CDM

    for k in range(0,CDM):
        sheet.cell(row=k+10, column=2).value = Muestras_list[k][0]
        sheet.cell(row=k+10, column=3).value = Muestras_list[k][1]
    
    salida.save(rutaOUT + "Para carga/Midiendo/"+ Empresa + '/'+File_name)
    salida.close

    os.remove(File_name)        
    
def EXCEL_visco(Empresa, CDM, Muestras_list, Fecha, Hmany ):
    file = os.path.dirname(os.path.realpath(__file__)) +"/Formularios/F-GA-8-3 Calculo de Viscosidad.xlsm"
    File_name = Fecha + '.xlsm'

    if not os.path.exists(rutaOUT + "Calculo de Viscosidad/Midiendo/"+ Empresa) == True:
        os.makedirs(rutaOUT + "Calculo de Viscosidad/Midiendo/"+ Empresa)
  
    shutil.copy(file, File_name)

    salida = load_workbook(File_name, keep_vba = True)
    sheet = salida.active

    sheet.cell(row=13,column=2).value = Empresa
    sheet.cell(row=13, column=6).value = Fecha
    sheet.cell(row=13, column=9).value = Hmany

    for k in range(0,Hmany):
        sheet.cell(row=k+15, column=2).value = Muestras_list[k][0]

    salida.save(rutaOUT + "Calculo de Viscosidad/Midiendo/"+ Empresa +'/'+File_name)
    salida.close

    os.remove(File_name)    
