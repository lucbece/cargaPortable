# -*- coding: utf-8 -*- 
#CAMBIOS : self.Empresas.SetItems(Global_Data.initialize()) en mainfram init
#         Empresa_INT = self.Empresas.GetCurrentSelection()
#Empresa = Global_Data.empresas_list[Empresa_INT] en data entry
import  wx
import GUI
import os
import Global_Data
import re
import shutil
from openpyxl import Workbook, load_workbook
import datetime
import copy


def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def date_converter(date):
    assert isinstance(date, wx.DateTime)
    if date.IsValid():
        ymd = map(int, date.FormatISODate().split('-'))
        return datetime.date(*ymd)
    else:
        return None
def EXCEL_carga(Empresa, CDM, Muestras_list, Fecha):
    if 'YPF' in Empresa:
        file = 'F-GA-9_2 Analisis de lubricantes_ypf.xlsx'
    else:
        file = "F-GA-9_2 Analisis de lubricantes.xlsx"

    File_name = Fecha + '.xlsx'    

    if not os.path.exists("X:/Para carga/Midiendo/"+ Empresa) == True:
        os.makedirs("X:/Para carga/Midiendo/"+ Empresa)
    shutil.copy(file, File_name)


    salida = load_workbook(File_name)
    sheet = salida.active

    sheet.cell(row=6,column=2).value = Empresa
    sheet.cell(row=6, column=5).value = Fecha
    sheet.cell(row=6, column=7).value = CDM

    for k in range(0,CDM):
        sheet.cell(row=k+10, column=2).value = Muestras_list[k][0]
        sheet.cell(row=k+10, column=3).value = Muestras_list[k][1]
    
    salida.save("X:/Para carga/Midiendo/"+ Empresa + '/'+File_name)
    salida.close      
    
def EXCEL_visco(Empresa, CDM, Muestras_list, Fecha, Hmany ):
    file = "F-GA-8-2 Calculo de Viscosidad con contador.xlsx"
    File_name = Fecha + '.xlsx'

    if not os.path.exists("C:/Calculo de Viscosidad/Midiendo/"+ Empresa) == True:
        os.makedirs("C:/Calculo de Viscosidad/Midiendo/"+ Empresa)
  
    shutil.copy(file, File_name)

    salida = load_workbook(File_name)
    sheet = salida.active

    sheet.cell(row=13,column=2).value = Empresa
    sheet.cell(row=13, column=5).value = Fecha
    sheet.cell(row=13, column=7).value = Hmany

    for k in range(0,Hmany):
        sheet.cell(row=k+15, column=2).value = Muestras_list[k][0]

    salida.save("C:/Calculo de Viscosidad/Midiendo/"+ Empresa +'/'+File_name)
    salida.close

    os.remove(File_name)    



class MainFrame(GUI.MainFrame):
    def __init__(self,parent):
        GUI.MainFrame.__init__(self,parent)
        self.Empresas.SetItems(Global_Data.initialize())

    def open_carga(self,event):
        os.startfile('X:/Para carga/Midiendo')

    def open_visco(self,event):
        os.startfile('X:/Calculo de Viscosidad/Midiendo')

    def data_entry(self,event):
        #Variables a usarse en la carga EXCEL
        RAW_Muestras = self.Muestras.GetValue().split('\n')
        #Agregar un check para ver si hay muestras repetidas

        for i in range(len(RAW_Muestras)):
            for j in range(len(RAW_Muestras)):
                if RAW_Muestras[i] == RAW_Muestras[j] and i != j :
                    errormsg.Show(True)
                    return
        for muestra in RAW_Muestras:
            if muestra == '':
                RAW_Muestras.remove(muestra)

        Empresa_INT = self.Empresas.GetCurrentSelection()
        Empresa = Global_Data.empresas_list[Empresa_INT]
        CDM = len(RAW_Muestras)

        Fecha = str(date_converter(self.date_picker.GetValue()))
        

        if self.TA_Simples.GetValue() == True :
            ORDERED_Muestras = copy.copy(RAW_Muestras)
            ORDERED_Muestras.sort(key=natural_keys)
            TA = self.TAS_Simples.GetValue()
            Muestras_list = []
            for i in range (0,CDM):
                Muestras_list.append([ORDERED_Muestras[i],TA])

            EXCEL_carga(Empresa, CDM, Muestras_list, Fecha)    

        elif self.TA_Completas.GetValue() == True:
            ORDERED_Muestras = copy.copy(RAW_Muestras)
            ORDERED_Muestras.sort(key=natural_keys)
            TA = self.TAS_Completas.GetValue()
            Muestras_list = []
            for i in range (0,CDM):
                Muestras_list.append([ORDERED_Muestras[i],TA])
            Hmany = CDM

            EXCEL_carga(Empresa, CDM, Muestras_list, Fecha)
            EXCEL_visco(Empresa, CDM, Muestras_list, Fecha, Hmany)        

        elif self.TA_Ambas.GetValue() == True :
            TAC = self.TAS_Completas.GetValue()
            TAS = self.TAS_Simples.GetValue()
            Rango_Completas = int(self.TA_Entry.GetValue())
            Muestras_Simples = []
            Muestras_Completas = []
            for i in range (0,Rango_Completas):
                Muestras_Completas.append(RAW_Muestras[i])
            for j in range (Rango_Completas,CDM):
                Muestras_Simples.append(RAW_Muestras[j])
            Muestras_Completas.sort(key=natural_keys)
            Muestras_Simples.sort(key=natural_keys)    

            Muestras_list = []
            for i in range (0,Rango_Completas):        
                Muestras_list.append([Muestras_Completas[i],TAC])
            for j in range (0,len(Muestras_Simples)):
                Muestras_list.append([Muestras_Simples[j],TAS])    
            

            EXCEL_carga(Empresa, CDM, Muestras_list, Fecha)
            EXCEL_visco(Empresa, CDM, Muestras_list, Fecha, Rango_Completas)

        int_list =Global_Data.numerar_list(RAW_Muestras,Muestras_list)
        nummsg.numeracion_muestras.Clear()
        nummsg.add_muestas(RAW_Muestras,int_list)
        nummsg.Show(True)

    def S_check_event(self,event):
        self.TA_Completas.SetValue(False)
        self.TA_Ambas.SetValue(False)

    def C_check_event(self,event):
        self.TA_Simples.SetValue(False)
        self.TA_Ambas.SetValue(False)

    def SyC_check_event(self,event):
        self.TA_Simples.SetValue(False)
        self.TA_Completas.SetValue(False)

    def agregar_equipo(self,event):
        muestra = self.muestras_pick.GetValue()
        self.Muestras.AppendText(muestra + '\n')            

    def clear_input(self,event):
        self.Muestras.Clear()
        self.TA_Entry.Clear()

    def select_database(self,event):
        self.empresas_int = self.Empresas.GetCurrentSelection()
        lista_muestras = Global_Data.pick_database(self.empresas_int)
        lista_muestras.sort(key = natural_keys)
        self.muestras_pick.Clear()

        for i in range (len(lista_muestras)):
            self.muestras_pick.Append(lista_muestras[i])

    def create_empresa(self,event):
        empresamsg.Show(True)

    def update_list(self):
        self.Empresas.SetItems(Global_Data.initialize())
                

class ErrorDiag(GUI.ErrorDiag):
    def __init__(self,parent):
        GUI.ErrorDiag.__init__(self,parent)
        
    def close_dialog(self,event):
        errormsg.Destroy()   

class new_empresa_dialog(GUI.new_empresa_dialog):
    def __init__(self,parent):
        GUI.new_empresa_dialog.__init__(self,parent) 

    def append_empresa(self,event):
        new_empresa =  self.ned_empresa.GetValue()
        Global_Data.add_empresa(new_empresa)
        frame.update_list()
        empresamsg.Show(False)

class numeracion_dialog(GUI.numeracion_dialog):
    def __init__(self,parent):
        GUI.numeracion_dialog.__init__(self,parent)

    def add_muestas(self,raw_list,int_list):
        for i in range(len(raw_list)):
            self.numeracion_muestras.AppendText(str(int_list[i]) + '-  ' + raw_list[i] + '\n')            



        


app = wx.App(False) 
frame = MainFrame(None) 
frame.Show(True) 
errormsg = ErrorDiag(None)
errormsg.Show(False)
empresamsg = new_empresa_dialog(None)
empresamsg.Show(False)
nummsg = numeracion_dialog(None)
nummsg.Show(False)
app.MainLoop() 

